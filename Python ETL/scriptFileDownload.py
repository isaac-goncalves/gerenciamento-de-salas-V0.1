import sys
import psycopg2
import pandas as pd
from datetime import datetime
import time
from openpyxl import Workbook, load_workbook

start_time = time.time()

if len(sys.argv) < 2:
    print("Please provide the file path as an argument.")
    sys.exit(1)

# Get the file path from the command-line argument
file_path = sys.argv[1]
course_id = int(sys.argv[2])
print(course_id)
# Establish a connection to your PostgreSQL database
connection = psycopg2.connect(
    #host='localhost',
    host = '192.168.0.53',
    port='5432',
    database='gerenciamento-de-salas',
    user='postgres',
    password='2406'
)

cursor = connection.cursor()

cursor.execute("SELECT name, id, surname, email FROM professores;")

professores = cursor.fetchall()

print(professores)


query = "SELECT name, id, surname, email FROM professores WHERE course_id =" + str(course_id) + ";" 
df = pd.read_sql_query(query, connection)

connection.close()

# Load the existing Excel file
workbook = load_workbook('BASE_DADOS_ADS-1.9 - Copia (2).xlsx')

sheet = workbook['Mock_Tables']  # Replace with your sheet name

# Define the starting row and column
start_row = 4
start_col = 5

# Exclude first ID column and last two timestamp columns
columns_to_insert = df.columns

# Insert DataFrame values into specific cells excluding selected columns
for _, row in df.iterrows():
    for idx, col in enumerate(columns_to_insert, start=start_col):
        sheet.cell(row=start_row, column=idx, value=row[col])
    start_row += 1



# Access the specific sheet in the workbook
# sheet = workbook['Mock_Tables']  # Replace with your sheet name

# # Define the starting row and column
# start_row = 3
# start_col = 4

# Insert DataFrame values into specific cells
#ignore first id and last two timestamps columns
# for r in df.iterrows():
#     for index, value in enumerate(r[1], start=0):
#         sheet.cell(row=start_row, column=start_col + index, value=value)
#     start_row += 1

randvaluefortimestamp = str(datetime.now().strftime("%Y%m%d%H%M%S"))



# Save the modified Excel file
workbook.save('BASE_DADOS_ADS-1.9 - Copia - Copia (3)'+randvaluefortimestamp+'.xlsx')
# df.to_excel(excel_writer, sheet_name='Mock_Tables', startrow=3, startcol=4, index=False)

# Write the data to an Excel file
# excel_writer.save()

